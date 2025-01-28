import io
import base64
import xlsxwriter
import openpyxl
from odoo import models, fields, api
from odoo.exceptions import UserError
import logging
_logger = logging.getLogger(__name__)

class InventorySession(models.Model):
    _name = 'inventory.session'
    _description = 'Inventory Session'

    name = fields.Char(string="Session Name", required=True, default="New")
    session_date = fields.Datetime(string="Session Date", default=fields.Datetime.now)
    group1_date_debut = fields.Date(string="Date Début Groupe 1")
    group1_date_fin = fields.Date(string="Date Fin Groupe 1")
    group2_date_debut = fields.Date(string="Date Début Groupe 2")
    group2_date_fin = fields.Date(string="Date Fin Groupe 2")
    session_line_ids = fields.One2many('inventory.session.line', 'session_id', string="Session Lines")
    state = fields.Selection([
        ('draft', 'Non validé'),
        ('partially_validated', 'Validé pour certain'),
        ('validated', 'Validé')
    ], string="État", default='draft', readonly=True)

    ###################################################

    def action_view_red_lines(self):
        """Open a tree view displaying only red lines."""
        self.ensure_one()
        _logger.info(f"Context passed: {self.env.context}")
        return {
            'type': 'ir.actions.act_window',
            'name': 'Red Lines',
            'res_model': 'inventory.session.line',
            'view_mode': 'tree',
            'domain': [('session_id', '=', self.id), ('difference', '!=', 0), ('final_qty', '=', 0)],
            'context': {'default_session_id': self.id},

        }

    #################################################

    def action_export_session_lines(self):
        """Export session lines for the current session to an Excel file."""

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Session Lines')


        header_format = workbook.add_format({'bold': True, 'bg_color': '#F9DA04', 'border': 1})
        data_format = workbook.add_format({'border': 1})


        headers = ['Emplacement', 'Produit', ' Quantité Group 1', ' Quantité Group 2']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)


        row = 1
        for line in self.session_line_ids:
            worksheet.write(row, 0, line.location_id.name or '', data_format)
            worksheet.write(row, 1, line.product_id.name or '', data_format)
            worksheet.write(row, 2, line.group_1_qty, data_format)
            worksheet.write(row, 3, line.group_2_qty, data_format)

            row += 1


        workbook.close()
        output.seek(0)


        attachment = self.env['ir.attachment'].create({
            'name': f'Session_Lines_{self.name}.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
        })


        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_import_session_lines(self):
        """Open a wizard to upload an Excel file."""
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Session Lines',
            'res_model': 'inventory.session.line.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_session_id': self.id},
        }

    def action_create_quant_lines(self):
        """Créer ou mettre à jour les lignes de stock.quant en fonction des lignes vertes uniquement et mettre à jour l'état."""
        self.ensure_one()
        all_green = True
        has_red = False


        self.session_line_ids._compute_difference()


        green_lines = self.session_line_ids.filtered(lambda line: line.difference == 0 and line.final_qty > 0)

        if not green_lines:
            raise UserError("No green lines found to process.")

        for line in green_lines:
            _logger.info(
                f"Processing green line: {line.product_id.name}, Final Qty: {line.final_qty}, Location: {line.location_id.name}"
            )

            # Update or create stock.quant
            quant = self.env['stock.quant'].search([
                ('location_id', '=', line.location_id.id),
                ('product_id', '=', line.product_id.id),
            ], limit=1)

            if quant:
                quant.write({'inventory_quantity': line.final_qty})
                _logger.info(f"Stock.quant updated: {quant.id} for Product: {line.product_id.name}")
            else:
                new_quant = self.env['stock.quant'].create({
                    'location_id': line.location_id.id,
                    'product_id': line.product_id.id,
                    'inventory_quantity': line.final_qty,
                })
                _logger.info(f"New stock.quant created: {new_quant.id} for Product: {line.product_id.name}")


        red_lines = self.session_line_ids.filtered(lambda line: line.difference != 0 )
        if red_lines:
            self.state = 'partially_validated'
            _logger.info("Session state set to: validé pour certain (partially validated)")
        else:
            self.state = 'validated'
            _logger.info("Session state set to: validé (validated)")


        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
            'params': {
                'display_notification': {
                    'title': 'Succès',
                    'message': 'Les lignes vertes ont été mises à jour dans stock.quant. État de la session mis à jour.',
                    'type': 'success',
                    'sticky': False,
                }
            },
        }

    ###################################################################
    def action_process_blue_lines(self):
        """Process blue lines (difference != 0 and final_qty > 0), update stock.quant, and review session state."""
        self.ensure_one()

        # Filter blue lines
        blue_lines = self.session_line_ids.filtered(lambda line: line.difference != 0 and line.final_qty > 0)

        if not blue_lines:
            raise UserError("No blue lines found to process.")

        for line in blue_lines:
            _logger.info(
                f"Processing blue line: {line.product_id.name}, Final Qty: {line.final_qty}, Location: {line.location_id.name}"
            )


            quant = self.env['stock.quant'].search([
                ('location_id', '=', line.location_id.id),
                ('product_id', '=', line.product_id.id),
            ], limit=1)

            if quant:
                quant.write({'inventory_quantity': line.final_qty})
                _logger.info(f"Stock.quant updated: {quant.id} for Product: {line.product_id.name}")
            else:
                new_quant = self.env['stock.quant'].create({
                    'location_id': line.location_id.id,
                    'product_id': line.product_id.id,
                    'inventory_quantity': line.final_qty,
                })
                _logger.info(f"New stock.quant created: {new_quant.id} for Product: {line.product_id.name}")


        red_lines = self.session_line_ids.filtered(lambda line: line.difference != 0 and line.final_qty == 0)
        if red_lines:
            self.state = 'partially_validated'
            _logger.info("Session state set to: validé pour certain (partially validated)")
        else:
            self.state = 'validated'
            _logger.info("Session state set to: validé (validated)")


        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
            'params': {
                'display_notification': {
                    'title': 'Succès',
                    'message': 'Les lignes corrigées ont été mises à jour dans stock.quant. État de la session mis à jour.',
                    'type': 'success',
                    'sticky': False,
                }
            },
        }

class InventorySessionLine(models.Model):
    _name = 'inventory.session.line'
    _description = 'Inventory Session Line'

    session_id = fields.Many2one('inventory.session', string="Session", ondelete="cascade")
    location_id = fields.Many2one('stock.location', string="Emplacement", required=True)
    product_id = fields.Many2one('product.product', string="Produit", required=True)
    group_1_qty = fields.Float(string="Quantité Groupe 1 ")
    group_2_qty = fields.Float(string="Quantité Groupe 2 ")
    difference = fields.Float(string="Écart", compute="_compute_difference", store=True)
    final_qty = fields.Float(string="Quantité Physique", store=True, default=0.0)



    @api.depends('group_1_qty', 'group_2_qty')
    def _compute_difference(self):
        for line in self:
            line.difference = line.group_1_qty - line.group_2_qty


    @api.onchange('difference')
    def _onchange_difference(self):
        for line in self:
            if line.difference == 0:
                line.final_qty = line.group_1_qty
            else:
                line.final_qty = 0.0









class InventorySessionLineImportWizard(models.TransientModel):
    _name = 'inventory.session.line.import.wizard'
    _description = 'Import Session Lines Wizard'

    session_id = fields.Many2one('inventory.session', string="Session", required=True)
    file = fields.Binary(string="File", required=True)
    file_name = fields.Char(string="File Name")

    def action_import_lines(self):
        """Process the uploaded file and create session lines."""
        if not self.file:
            raise UserError("Please upload a file.")


        file_data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(io.BytesIO(file_data))
        sheet = workbook.active


        session_lines = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
            location_name, product_name, group_1_qty, group_2_qty = row[:4]


            location = self.env['stock.location'].search([('name', '=', location_name)], limit=1)
            product = self.env['product.product'].search([('name', '=', product_name)], limit=1)

            if not location or not product:
                raise UserError(f"Invalid location or product in row: {row}")


            session_lines.append({
                'session_id': self.session_id.id,
                'location_id': location.id,
                'product_id': product.id,
                'group_1_qty': group_1_qty or 0.0,
                'group_2_qty': group_2_qty or 0.0,
            })


        lines = self.env['inventory.session.line'].create(session_lines)


        for line in lines:
            line._compute_difference()
            line._onchange_difference()

        return {'type': 'ir.actions.act_window_close'}
